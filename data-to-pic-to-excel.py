import code128
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.drawing.image import Image as xlImage
import os

# Get barcode value
barcode_param = 'ATL-003111'

# Create barcode image
barcode_image = code128.image(barcode_param, height=120)

# Create empty image for barcode + text
top_bott_margin = 60
l_r_margin = 0
new_height = barcode_image.height + (2 * top_bott_margin)
new_width = barcode_image.width + (2 * l_r_margin)
new_image = Image.new('RGB', (new_width, new_height), (255, 255, 255))

# put barcode on new image
barcode_y = 50
new_image.paste(barcode_image, (0, barcode_y))

# object to draw text
draw = ImageDraw.Draw(new_image)

# Define custom text size and font
h1_size = 36

h1_font = ImageFont.truetype(
    "C:\\Windows\\Fonts\\Calibri\\Calibri\\calibri.ttf", h1_size)
# Define custom text
center_barcode_value = (barcode_image.width / 2) - len(barcode_param) * 8

# Draw text on picture
draw.text((center_barcode_value, (new_height - h1_size - 15)),
          barcode_param, fill=(0, 0, 0), font=h1_font)

# save in file
new_image.save('barcode_image.png', 'PNG')

#-------------------------------------------------------------------------------------------------
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

#change to the location and name of your image of the barcode
png_loc = r'./barcode_image.png'
FILENAME = 'test'                              # REPLACE WITH YOUR FILENAME
COPY = 'test1'                           # REPLACE WITH YOUR FILENAME

try:
    wb = load_workbook(FILENAME + '.xlsx')                 # REPLACE WITH YOUR FILENAME
except FileNotFoundError:
    print("File not found, creating new file")
    wb = openpyxl.Workbook()
    wb.save(FILENAME + '.xlsx')                            # REPLACE WITH YOUR FILENAME

ws = wb.active
my_png = Image(png_loc)

# scaling the image to 32% height and 44% width of its original size
my_png.height = 0.3228 * my_png.height
my_png.width = 0.4458 * my_png.width

# add and scale the image to the cell
ws.add_image(my_png, 'BQ6')


wb.save(COPY + '.xlsx')

# Adapted from https://stackoverflow.com/questions/10888969/insert-image-in-openpyxl

#-------------------------------------------------------------------------------------------------
import win32com.client
from pywintypes import com_error

# Path to original excel file
rel_xl_path = COPY + ".xlsx"
abs_xl_path = os.path.abspath(rel_xl_path)
WB_PATH = abs_xl_path

# PDF path when saving
PATH_TO_PDF = abs_xl_path.replace('.xlsx', '.pdf')
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
try:
    print('Start conversion to PDF')
    # Open
    wb = excel.Workbooks.Open(WB_PATH)
    # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
    ws_index_list = [1]
    wb.WorkSheets(ws_index_list).Select()
    # Save
    wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
except com_error as e:
    print('failed.')
else:
    print('Succeeded.')
finally:
    wb.Close()
    excel.Quit()
