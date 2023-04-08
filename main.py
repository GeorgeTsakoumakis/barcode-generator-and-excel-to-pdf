FILENAME = input('Enter name of Excel file to write to - DO NOT include extention (make sure it is in same directory as exe: \n>')                              # REPLACE WITH YOUR FILENAME
COPY = input('Enter name of copy of Excel file to be created and exported later - DO NOT include extention: \n>')

def barcode_generator(data):
    import code128
    from PIL import Image, ImageDraw, ImageFont
    from openpyxl.drawing.image import Image as xlImage
    import os

    # Get barcode value
    barcode_param = data             # REPLACE WITH YOUR BARCODE VALUE

    # Create barcode image
    barcode_image = code128.image(barcode_param, height=120)

    # Create empty image for barcode + text
    top_bott_margin = 60
    l_r_margin = 0
    new_height = barcode_image.height + (2 * top_bott_margin)
    new_width = barcode_image.width + (2 * l_r_margin)
    new_image = Image.new('RGB', (new_width, new_height), (255, 255, 255))

    # put barcode on new image
    barcode_y = 50
    new_image.paste(barcode_image, (0, barcode_y))

    # object to draw text
    draw = ImageDraw.Draw(new_image)

    # Define custom text size and font
    h1_size = 36

    h1_font = ImageFont.truetype(
        "C:\\Windows\\Fonts\\Calibri\\Calibri\\calibri.ttf", h1_size)
    # Define custom text
    center_barcode_value = (barcode_image.width / 2) - len(barcode_param) * 8

    # Draw text on picture
    draw.text((center_barcode_value, (new_height - h1_size - 15)),
            barcode_param, fill=(0, 0, 0), font=h1_font)

    # save in file
    if not os.path.exists('image'):
        os.makedirs('image')
    try:
        new_image.save('image/barcode_image_'+data+'.png', 'PNG')
    except FileNotFoundError:
        print("File not found, creating new file")
        quit()

#-------------------------------------------------------------------------------------------------
def pic_in_excel(data):
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image

    #change to the location and name of your image of the barcode
    png_loc = r'image/barcode_image_'+data+'.png'


    try:
        wb = load_workbook(FILENAME + '.xlsx')                
    except FileNotFoundError:
        print("File not found, creating new file")
        wb = openpyxl.Workbook()
        wb.save(FILENAME + '.xlsx')                           

    ws = wb.active
    my_png = Image(png_loc)

    # scaling the image to 32% height and 44% width of its original size
    my_png.height = 0.3228 * my_png.height
    my_png.width = 0.4458 * my_png.width

    # add and scale the image to the cell
    cell = input('Enter cell to place barcode in: \n>')
    ws.add_image(my_png, cell)                 


    wb.save(COPY + '.xlsx')

    # Adapted from https://stackoverflow.com/questions/10888969/insert-image-in-openpyxl

#-------------------------------------------------------------------------------------------------
def xl_to_pdf(data):
    import win32com.client
    from pywintypes import com_error
    import os


    # Path to original excel file
    rel_xl_path = COPY + ".xlsx" # relative path to excel file
    abs_xl_path = os.path.abspath(rel_xl_path) 
    WB_PATH = abs_xl_path # absolute path to excel file

    # PDF path when saving
    PATH_TO_PDF = WB_PATH.replace(COPY + '.xlsx', 'PDFs\\'+COPY+'.pdf') # absolute path to pdf file in PDFs folder
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        print('Start conversion to PDF')
        # Open
        wb = excel.Workbooks.Open(WB_PATH)
        # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
        ws_index_list = [1]
        wb.WorkSheets(ws_index_list).Select()
        # Save
        output_filename = data + ".pdf"
        PATH_TO_PDF = PATH_TO_PDF.replace(COPY + '.pdf', output_filename)
        if not os.path.exists('PDFs'):
            os.makedirs('PDFs') # create PDFs folder if it doesn't exist
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except com_error as e:
        print('failed.')
    else:
        print('Succeeded.')
    finally:
        wb.Close()
        excel.Quit()

def main(data):
    if data == '':
        print('No data')
        return
    barcode_generator(data)
    pic_in_excel(data)
    xl_to_pdf(data)

try:
    data = input('Enter barcode data: ')
    main(data)
except TypeError:
    print('No data')